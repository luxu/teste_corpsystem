import os

import openpyxl
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from . import serializers
from .models import Venda


@api_view(['GET'])
def vendas(request):
    vendas = Venda.objects.all()
    serializer = serializers.VendaSerializer(vendas, many=True)
    return Response(
        serializer.data,
        status=status.HTTP_200_OK
    )


@api_view(['POST'])
def criar_venda(request):
    serializer = serializers.VendaCriarSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
def atualizar_venda(request, id):
    if id is None:
        return Response(status=status.HTTP_400_BAD_REQUEST)
    try:
        venda = Venda.objects.get(id=id)
        serializer = serializers.VendaAtualizarSerializer(
            venda, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Venda.DoesNotExist:
        result = Response(
            f"A venda com ID:{id} não existe!",
            status=status.HTTP_404_NOT_FOUND
        )
        return result


@api_view(['DELETE'])
def deletar_venda(request, id):
    if id is None:
        return Response(status=status.HTTP_400_BAD_REQUEST)
    try:
        venda = Venda.objects.get(id=id)
        venda.delete()
        return Response(
            f"A venda..:{venda.id} foi deletada com sucesso!",
            status=status.HTTP_204_NO_CONTENT
        )
    except Venda.DoesNotExist:
        result = Response(
            f"A venda com ID:{id} não existe!",
            status=status.HTTP_404_NOT_FOUND
        )
        return result


@api_view(['GET'])
def vendas_efetuadas(request):
    # A lógica é gerar as VENDAS que tenham sido pagas, ou seja, tiver vendas em aberto NÃO entra no relatório
    vendas = Venda.objects.filter(total_a_pagar__gt=0)
    if data_venda := request.GET.get('data_venda'):
        vendas = vendas.filter(data_venda__gt=data_venda)
    if vendedor := request.GET.get('vendedor'):
        vendas = vendas.filter(vendedor=vendedor)
    if cliente := request.GET.get('cliente'):
        vendas = vendas.filter(cliente=cliente)

    nome_arquivo_excel = gerar_nota_fiscal_excel(vendas)
    with open(nome_arquivo_excel, 'rb') as excel:
        response = HttpResponse(
            excel.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(nome_arquivo_excel)}'

    nome_arquivo_pdf = gerar_nota_fiscal_pdf(vendas)
    with open(nome_arquivo_pdf, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(nome_arquivo_pdf)}'
        return response


def gerar_nota_fiscal_excel(vendas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nota Fiscal"
    nro_nota = 0

    linha_inicial = 1

    for venda in vendas:
        nro_nota = venda.id
        ws.merge_cells(start_row=linha_inicial, start_column=1, end_row=linha_inicial, end_column=5)
        ws.cell(row=linha_inicial, column=1, value=f"Nota Fiscal - Venda ID: {venda.id}").font = Font(bold=True)
        linha_inicial += 1

        ws.cell(row=linha_inicial, column=1, value="Cliente:")
        ws.cell(row=linha_inicial, column=2, value=venda.cliente.nome)
        linha_inicial += 1

        ws.cell(row=linha_inicial, column=1, value="Vendedor:")
        ws.cell(row=linha_inicial, column=2, value=venda.vendedor.nome)
        linha_inicial += 1

        ws.cell(row=linha_inicial, column=1, value="Data da Venda:")
        ws.cell(row=linha_inicial, column=2, value=venda.data_venda.strftime('%Y-%m-%d %H:%M:%S'))
        linha_inicial += 2

        ws.cell(row=linha_inicial, column=1, value="Produto").font = Font(bold=True)
        ws.cell(row=linha_inicial, column=2, value="Quantidade").font = Font(bold=True)
        ws.cell(row=linha_inicial, column=3, value="Preço Unitário").font = Font(bold=True)
        ws.cell(row=linha_inicial, column=4, value="Total").font = Font(bold=True)
        linha_inicial += 1

        total_venda = 0
        for itens_vendas in venda.vendas.all():
            ws.cell(row=linha_inicial, column=1, value=itens_vendas.produto.nome)
            ws.cell(row=linha_inicial, column=2, value=itens_vendas.quantidade)
            ws.cell(row=linha_inicial, column=3,
                    value=f"R$ {itens_vendas.produto.preco:,.2f}".replace('.', ',').replace(',', '.', 1))
            total_item = itens_vendas.quantidade * itens_vendas.produto.preco
            ws.cell(row=linha_inicial, column=4, value=f"R$ {total_item:,.2f}".replace('.', ',').replace(',', '.', 1))
            total_venda += total_item
            linha_inicial += 1

        ws.merge_cells(start_row=linha_inicial, start_column=1, end_row=linha_inicial, end_column=3)
        ws.cell(row=linha_inicial, column=1, value="Total da Venda:").font = Font(bold=True)
        ws.cell(row=linha_inicial, column=4,
                value=f"R$ {total_venda:,.2f}".replace('.', ',').replace(',', '.', 1)).font = Font(bold=True)
        linha_inicial += 2


    nome_arquivo = f'media/NF_nro_{nro_nota}.xlsx'
    wb.save(nome_arquivo)

    return nome_arquivo


def gerar_nota_fiscal_pdf(vendas):
    # nome_arquivo = f'media/NF_nro_{vendas.values("id").first()}.pdf'
    nome_arquivo = 'media/nota_fiscal_vendas.pdf'
    pdf = SimpleDocTemplate(nome_arquivo, pagesize=A4)
    elementos = []

    styles = getSampleStyleSheet()
    estilo_titulo = styles['Title']
    estilo_cabecalho = styles['Heading2']
    estilo_normal = styles['Normal']

    for venda in vendas:
        elementos.append(Paragraph(f"Nota Fiscal - Venda ID: {venda.id}", estilo_titulo))
        elementos.append(Paragraph(f"Cliente: {venda.cliente}", estilo_cabecalho))
        elementos.append(Paragraph(f"Vendedor: {venda.vendedor.nome}", estilo_cabecalho))
        elementos.append(
            Paragraph(f"Data da Venda: {venda.data_venda.strftime('%Y-%m-%d %H:%M:%S')}", estilo_cabecalho))
        elementos.append(Paragraph(" ", estilo_normal))
        tabela_dados = [
            ['Produto', 'Quantidade', 'Preço Unitário', 'Total']
        ]
        total_venda = 0
        for itens_vendas in venda.vendas.all():
            preco_unitario_formatado = f"R$ {itens_vendas.produto.preco:,.2f}".replace('.', ',').replace(',', '.', 1)
            total_item = itens_vendas.quantidade * itens_vendas.produto.preco
            total_item_formatado = f"R$ {total_item:,.2f}".replace('.', ',').replace(',', '.', 1)
            tabela_dados.append([itens_vendas.produto, itens_vendas.quantidade, preco_unitario_formatado, total_item_formatado])
            total_venda += total_item

        tabela = Table(tabela_dados, colWidths=[8 * cm, 4 * cm, 4 * cm, 4 * cm])
        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabela)

        elementos.append(Paragraph(" ", estilo_normal))

        total_venda_formatado = f"R$ {total_venda:,.2f}".replace('.', ',').replace(',', '.', 1)
        elementos.append(Paragraph(f"Total da Venda: {total_venda_formatado}", estilo_cabecalho))

        elementos.append(Paragraph(" ", estilo_normal))

    pdf.build(elementos)
    return nome_arquivo
